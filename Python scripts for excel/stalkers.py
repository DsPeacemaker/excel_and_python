import stalker
import openpyxl as pyxl
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference

wb = pyxl.load_workbook("example.xlsx")
sheet = wb["stalkers"]

sheet.append(["Имя", "Ранг", "Опыт", "Клан", "Оружие", "Пистолет", "Комбинезон"])

status = {"Монолит": "00FF0000", "Наёмники": "00FF0000", "Бандиты": "00FF0000", "Нейтрал": "00FFFF00",
          "Долг": "00FFFF00", "Свобода": "0000FF00"}

dict_population = {"Монолит": 0, "Наёмники": 0, "Бандиты": 0, "Нейтрал": 0,
          "Долг": 0, "Свобода": 0}

for i in range(2, 2000):
    npc = stalker.Stalker()
    sheet[f"A{i}"] = npc.name
    sheet[f"B{i}"] = npc.rank
    sheet[f"C{i}"] = npc.rank_point
    sheet[f"D{i}"] = npc.clan
    sheet[f"E{i}"] = npc.first_weapon
    sheet[f"F{i}"] = npc.second_weapon
    sheet[f"G{i}"] = npc.armor

# color_back = PatternFill(fgColor="F0FA0AFC")
# diff_style = DifferentialStyle(fill=color_back)
# rule = Rule(type="expression", dxf=diff_style)
# rule.formula = ["$С2>930"]
# sheet.conditional_formatting.add("A2:G1999", rule)

ekzo = 0
name_list = []
for i in range(2, 2000):
    if sheet.cell(i, 4).value in status.keys():
        sheet.cell(i, 4).fill = PatternFill('solid', fgColor=status[sheet.cell(i, 4).value])

    if sheet.cell(i, 4).value in dict_population.keys():
        dict_population[sheet.cell(i, 4).value] = dict_population.get(sheet.cell(i, 4).value) + 1

    if sheet.cell(i, 7).value == 'Экзоскелет':
        ekzo += 1

    if sheet["K1"].value in sheet.cell(i, 1).value:
        name_list.append(sheet.cell(i, 1).value)

for i in range(1, 7):
    clans = ["Наёмники", "Монолит", "Нейтрал", "Долг", "Свобода", "Бандиты"]
    sheet[f"H{i}"] = clans[i-1]
    sheet[f"I{i}"] = dict_population[clans[i-1]]

# for i in range(1, 5):
#     rank = ["Новичек", "Опытный", "Ветеран", "Мастер"]
#     sheet[f"H{i+8}"] = rank[i-1]
#     sheet[f"I{i+8}"] = "=COUNTIF($B$2:$B$1999;H{i+8})"


sheet["H7"] = "Всего"
sheet["I7"] = "=SUM(I1:I6)"
sheet["I18"] = ekzo

for i, name in enumerate(name_list, start=1):
    sheet[f"L{i}"] = name
sheet["K2"] = len(name_list)

# Диаграмма
# chart = BarChart()
# data = Reference(worksheet=sheet["H1:I6"],
#                  min_row=1,
#                  max_row=6,
#                  min_col=8,
#                  max_col=9)
#
# chart.add_data(data, from_rows=True)
# chart.style = 16
# sheet.add_chart(chart, "U2")

wb.save('example.xlsx')
