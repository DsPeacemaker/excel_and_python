import openpyxl as pyxl

wb = pyxl.load_workbook('orders.xlsx', data_only=True)

# узнать имена всех "листов" файла
sheets = wb.sheetnames

# выбрать активный лист
# если нужно выбрать лист по имени пишем sheet = wb['name_sheet']
sheet = wb.active

# нашли нужные столбцы столбец В - артикул, стобец L - категория товара
# так же можно обращаться к столбцам с помощью sheet.cell(index_row, index_column)
# print(sheet.cell(row=7, column=2).value, sheet.cell(row=7, column=12).value)
art = sheet[f'B{7}'].value
category = sheet['L7'].value

# узнаем номер последней строки
max_row = sheet.max_row
dict_categories = {}

# строим соотношение категорий с артикулами
for i in range(7, max_row+1):
    art = sheet[f'B{i}'].value
    category = sheet[f'L{i}'].value

    if not art:
        continue

    if category not in dict_categories:
        dict_categories[category] = [art]
    else:
        dict_categories[category].append(art)

# Сортируем словарь
sorted_dict = dict(sorted(dict_categories.items()))

# Записываем в файл
with open('category.txt', 'w') as file:
    for key, value in sorted_dict.items():
        str_values = ', '.join(value)
        str_to_write = key + ' ' + str_values + '\n'
        file.write(str_to_write)
