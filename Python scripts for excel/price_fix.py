import openpyxl as pyxl
from openpyxl.styles import Font, PatternFill
import random

wb = pyxl.load_workbook('example.xlsx', data_only=True)
sheet = wb["Price"]

names_columns = {"A": 'Наименование', "B": 'Цена', "C": 'Продано', "D": 'Выручка', "E": "Остаток",
                 "G": "Категория", "H": "Выручка Категории"}
products = ['Картофель', "Арбуз", "Дыня", "Яблоки", "Апельсины", "Бананы", "Морковь", "Ананас", "Лук", "Виноград",
            "Гранат", "Капуста", "Свекла", "Зелень"]
dict_total = {}
correct_price_up = {'Гранат': 50, 'Ананас': 60, 'Бананы': 55}
correct_price_down = {'Апельсины': 20, 'Капуста': 15, 'Морковь': 13}

len_products = len(products)-1
sheet.append(names_columns)

for i in range(2, 350):
    sheet.cell(i, 1).value = products[random.randint(0, len_products)]
    sheet.cell(i, 2).value = random.randint(1, 40)
    sheet.cell(i, 3).value = random.randint(0, 100)
    sheet.cell(i, 4).value = f"=ROUND(B{i}*C{i}, 2)"
    sheet.cell(i, 5).value = random.randint(0, 20)

for i in range(2, 350):
    if sheet.cell(i, 1).value not in dict_total.keys():
        dict_total.setdefault(sheet.cell(i, 1).value, sheet.cell(i, 2).value * sheet.cell(i, 3).value)
    else:
        dict_total[sheet.cell(i, 1).value] += sheet.cell(i, 2).value * sheet.cell(i, 3).value

    if sheet.cell(i, 1).value in correct_price_up.keys():
        sheet.cell(i, 2).value = correct_price_up[sheet.cell(i, 1).value]
        sheet.cell(i, 2).font = Font(color="FF0000")

    if sheet.cell(i, 1).value in correct_price_down.keys():
        sheet.cell(i, 2).value = correct_price_down[sheet.cell(i, 1).value]
        sheet.cell(i, 2).font = Font(color="0000FF00")

    if sheet.cell(i, 5).value < 10:
        sheet.cell(i, 5).fill = PatternFill('solid', fgColor="FF0000")

    sheet.cell(i, 4).fill = PatternFill('solid', fgColor="0000FF00")

max_rows = sheet.max_row
sheet["F1"] = "Общая сумма"
sheet["F2"] = f"=SUM(D1:D{max_rows})"

for i, items in enumerate(dict_total.items(), start=2):
    sheet.cell(i, 7).value = items[0]
    sheet.cell(i, 8).value = items[1]

wb.save('example.xlsx')


